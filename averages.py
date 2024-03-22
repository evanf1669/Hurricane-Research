import openpyxl
from openpyxl import Workbook, load_workbook

#function for getting the average of each row 
def getaverage (file, sh = ''):
    book = load_workbook(file)
    
    
    if sh != '':
        sheet = book[sh]
    else:
        sheet = book.active

    dictionary = {}
    
    

    for row  in range(2, sheet.max_row+1):
        cell = 'A' + str(row)
        year = sheet[cell].value
        total = 0.0
        row_length = sheet[row]
        
        for col in range(2, sheet.max_column+1):

            total = total + float(sheet.cell(row, col).value)
            
        total = total / (len(row_length) - 1)
        dictionary[year] = float("%.4f" %total)
        
        
    return(dictionary)

def write(dict, finalCol, colName):
    book = load_workbook("Everything_Bagel.xlsx")
    sheet = book.active
    cell_temp = finalCol + "1"
    sheet[cell_temp] = colName
    for i, (key, value) in enumerate(dict.items()):
        if int(key) < 1900:
            continue
        for row  in range(2, sheet.max_row + 1):
            cell = finalCol + str(row)
            if int(sheet.cell(row, 1).value) == key:
                sheet[cell] = value
    book.save("Everything_Bagel.xlsx")
            


co2_mlo = getaverage("co2_mlo.xlsx")

write(co2_mlo, "B", "CO2")


darwin_originalData = getaverage("Darwin.xlsx", "Original Data")
darwin_anomaly = getaverage("Darwin.xlsx", "Anomaly")
darwin_standardizedData = getaverage("Darwin.xlsx", "Standardized Data")
heat_content_index_130 = getaverage("heat_content_index.xlsx", "130E-80W")
heat_content_index_160 = getaverage("heat_content_index.xlsx", "160E-80W")
heat_content_index_180 = getaverage("heat_content_index.xlsx", "180W-100W")
norm_nao = getaverage("norm_nao_monthly_b5001_current.xlsx")
reqsoi = getaverage("reqsoi.for.xlsx")
soi_anomaly = getaverage("soi.xlsx", "Anomaly")
soi_standardizedData = getaverage("soi.xlsx", "Standardized Data")
ssoi_atl_natl =getaverage("sstoi.atl.indicies.xlsx", "NATL")
ssoi_atl_natla =getaverage("sstoi.atl.indicies.xlsx", "ANOM NATL")
ssoi_atl_satl =getaverage("sstoi.atl.indicies.xlsx", "SATL")
ssoi_atl_satla =getaverage("sstoi.atl.indicies.xlsx", "ANOM SATL")
ssoi_atl_trop =getaverage("sstoi.atl.indicies.xlsx", "TROP")
ssoi_atl_tropa =getaverage("sstoi.atl.indicies.xlsx", "ANOM TROP")
sstoi_nino12 = getaverage("sstoi.indicies.xlsx", "NINO1+2")
sstoi_nino12a = getaverage("sstoi.indicies.xlsx", "ANOM NINO1+2")
sstoi_nino3 = getaverage("sstoi.indicies.xlsx", "NINO3")
sstoi_nino3a = getaverage("sstoi.indicies.xlsx", "ANOM NINO3")
sstoi_nino4 = getaverage("sstoi.indicies.xlsx", "NINO4")
sstoi_nino4a = getaverage("sstoi.indicies.xlsx", "ANOM NINO4")
sstoi_nino34 = getaverage("sstoi.indicies.xlsx", "NINO3.4")
sstoi_nino34a = getaverage("sstoi.indicies.xlsx", "ANOM NINO3.4")
tahiti_sea_level_press = getaverage("tahiti.xlsx", "sea level press (1000mb subtr)")
tahiti__sea_level_stnd = getaverage("tahiti.xlsx", "sea level stnd.")
tahiti_anomaly = getaverage("tahiti.xlsx", "anomaly")
wksst_nino1 = getaverage("wksst9120_for.xlsx", "Nino1+2SST")
wksst_nino1a = getaverage("wksst9120_for.xlsx", "Nino1+2SSTA")
wksst_nino3 = getaverage("wksst9120_for.xlsx", "Nino3SST")
wksst_nino3a = getaverage("wksst9120_for.xlsx", "Nino3SSTA")
wksst_nino34 = getaverage("wksst9120_for.xlsx", "Nino34SST")
wksst_nino34a = getaverage("wksst9120_for.xlsx", "Nino34SSTA")
wksst_nino4 = getaverage("wksst9120_for.xlsx", "Nino4SST")
wksst_nino4a = getaverage("wksst9120_for.xlsx", "Nino4SSTA")
z500t_originalData = getaverage("z500t.xlsx", "Original Data")
z500t_anomaly = getaverage("z500t.xlsx", "Anomaly")
z500t_standardizedData = getaverage("z500t.xlsx", "Standardized Data")

#list = [co2_mlo, darwin_originalData, darwin_anomaly, darwin_standardizedData, heat_content_index_130, heat_content_index_160, heat_content_index_180,
#norm_nao, reqsoi, soi_anomaly, soi_standardizedData,ssoi_atl_natl, ssoi_atl_natla, ssoi_atl_satl, ssoi_atl_satla, ssoi_atl_trop, ssoi_atl_tropa,
#sstoi_nino12, sstoi_nino12a, sstoi_nino3, sstoi_nino3a, sstoi_nino34, sstoi_nino34a, sstoi_nino4, sstoi_nino4a, tahiti_sea_level_press, tahiti_anomaly, tahiti__sea_level_stnd,
#wksst_nino1, wksst_nino1a, wksst_nino3, wksst_nino3a, wksst_nino34, wksst_nino34a, wksst_nino4, wksst_nino4a, z500t_originalData, z500t_anomaly, z500t_standardizedData
#]

write(darwin_originalData, "C", "Darwin Original Data")
write(darwin_anomaly, "D", "Darwin Anomaly")
write(darwin_standardizedData, "E", "Darwin Standardized Data")
write(heat_content_index_130, "F", "130E-80W HCI")
write(heat_content_index_160, "G", "160E-80W HCI")
write(heat_content_index_180, "H", "180W-100W HCI")
write(norm_nao, "I", "norm_nao")
write(reqsoi, "J", "reqsoi")
write(soi_anomaly, "K", "soi anomaly")
write(soi_standardizedData, "L", "soi standardized data")
write(ssoi_atl_natl, "M", "NATL")
write(ssoi_atl_natla, "N", "NATLa")
write(ssoi_atl_satl, "O", "SATL")
write(ssoi_atl_satla, "P", "SATLa")
write(ssoi_atl_trop, "Q", "TROP")
write(ssoi_atl_tropa, "R", "TROPa")
write(sstoi_nino12, "S", "NINO1+2")
write(sstoi_nino12a, "T", "NINO 1+2a")
write(sstoi_nino3, "U", "NINO3")
write(sstoi_nino3a, "V", "NINO3a")
write(sstoi_nino4, "W", "NINO4")
write(sstoi_nino4a, "X", "NINO4a")
write(sstoi_nino34, "Y", "NINO3.4")
write(sstoi_nino34a, "Z", "NINO3.4a")
write(tahiti_sea_level_press, "AA", "sea level press (1000mb subtr)")
write(tahiti__sea_level_stnd, "AB", "sea level stnd")
write(tahiti_anomaly, "AC", "tahiti anomaly")
write(wksst_nino1, "AD", "Nino1+2SST")
write(wksst_nino1a, "AE", "Nino1+2SSTA")
write(wksst_nino3, "AF", "Nino3SST")
write(wksst_nino3a, "AG", "Nino3SSTA")
write(wksst_nino34, "AH", "Nino34SST")
write(wksst_nino34a, "AI", "Nino34SSTA")
write(wksst_nino4, "AJ", "Nino4SST")
write(wksst_nino4a, "AK", "Nino4SSTA")
write(z500t_originalData, "AL", "z500t original data")
write(z500t_anomaly, "AM", "z500t anomaly")
write(z500t_standardizedData, "AN", "z500t standardized data")
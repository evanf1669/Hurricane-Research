import openpyxl
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt

book = load_workbook("Everything_Bagel.xlsx")
sheet = book.active

#Makes a graph with each combination of column, only graphs with significant descriptive statistics are kept

x_axis = []
y_axis = []

for col in range(2, sheet.max_column + 1):
    for col2 in range(2, sheet.max_column + 1):
        for row in range(2, sheet.max_row + 1):


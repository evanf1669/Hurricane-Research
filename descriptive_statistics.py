import openpyxl
from openpyxl import Workbook, load_workbook
import math
import matplotlib.pyplot as plt
import scipy.stats as ss

book = load_workbook("Everything_Bagel.xlsx")
sheet = book.active


def getCC(col1, col2):
    x = 0
    y = 0
    x2 = 0
    y2 = 0
    xy = 0
    sampleSize = 0
    for row in range(2, sheet.max_row + 1):

        if str(sheet.cell(row, col1).value) == "None" or str(sheet.cell(row, col2).value) == "None" or str(sheet.cell(row, col1).value) == "*" or str(sheet.cell(row, col2).value) == "*" or str(sheet.cell(row, col1).value) == "NA" or str(sheet.cell(row, col2).value) == "NA":
            continue
        else:
            x += float(sheet.cell(row, col1).value)
            x2 += math.pow(float(sheet.cell(row, col1).value), 2)
            xy += (float(sheet.cell(row, col1).value)) * (float(sheet.cell(row, col2).value))
            y += float(sheet.cell(row, col2).value)
            y2 += math.pow(float(sheet.cell(row, col2).value), 2)
            sampleSize += 1
    
    CC = ((sampleSize * xy) - (x * y)) / math.sqrt( ( (sampleSize * x2) - math.pow(x, 2) ) * ( (sampleSize * y2) - math.pow(y, 2) ) )
    return CC

def graph(col1, col2):
    x_axis = []
    y_axis = []
    title = str(sheet.cell(1, col1).value) + " vs " + str(sheet.cell(1, col2).value)
    
    for row in range(2, sheet.max_row + 1):

        if str(sheet.cell(row, col1).value) == "None" or str(sheet.cell(row, col2).value) == "None" or str(sheet.cell(row, col1).value) == "*" or str(sheet.cell(row, col2).value) == "*" or str(sheet.cell(row, col1).value) == "NA" or str(sheet.cell(row, col2).value) == "NA":
            continue
        else: 
            x_axis.append(float(sheet.cell(row, col1).value))
            y_axis.append(float(sheet.cell(row, col2).value))

    plt.scatter(x_axis, y_axis, color = "red")
    plt.title(title)
    plt.xlabel(str(sheet.cell(1, col1).value))
    plt.ylabel(str(sheet.cell(1, col2).value))
    plt.grid(True)
    plt.show()

def getSamplesize(col1, col2):
    sampleSize = 0.0
    for row in range(2, sheet.max_row + 1):

        if str(sheet.cell(row, col1).value) == "None" or str(sheet.cell(row, col2).value) == "None" or str(sheet.cell(row, col1).value) == "*" or str(sheet.cell(row, col2).value) == "*" or str(sheet.cell(row, col1).value) == "NA" or str(sheet.cell(row, col2).value) == "NA":
            continue
        else:
            sampleSize += 1.0
    return sampleSize

def Tstat(col1, col2):
    p = getCC(col1, col2)
    n = getSamplesize(col1, col2)

    tScore = p * math.sqrt((n - 2) / (1 - math.pow(p, 2)))
    return tScore
    
def ppvalue (col1, col2):
    x_axis = []
    y_axis = []
    title = str(sheet.cell(1, col1).value) + " vs " + str(sheet.cell(1, col2).value)
    
    for row in range(2, sheet.max_row + 1):

        if str(sheet.cell(row, col1).value) == "None" or str(sheet.cell(row, col2).value) == "None" or str(sheet.cell(row, col1).value) == "*" or str(sheet.cell(row, col2).value) == "*" or str(sheet.cell(row, col1).value) == "NA" or str(sheet.cell(row, col2).value) == "NA":
            continue
        else: 
            x_axis.append(float(sheet.cell(row, col1).value))
            y_axis.append(float(sheet.cell(row, col2).value))
    
    return ss.pearsonr(x_axis, y_axis).pvalue

def spvalue (col1, col2):
    x_axis = []
    y_axis = []
    title = str(sheet.cell(1, col1).value) + " vs " + str(sheet.cell(1, col2).value)
    
    for row in range(2, sheet.max_row + 1):

        if str(sheet.cell(row, col1).value) == "None" or str(sheet.cell(row, col2).value) == "None" or str(sheet.cell(row, col1).value) == "*" or str(sheet.cell(row, col2).value) == "*" or str(sheet.cell(row, col1).value) == "NA" or str(sheet.cell(row, col2).value) == "NA":
            continue
        else: 
            x_axis.append(float(sheet.cell(row, col1).value))
            y_axis.append(float(sheet.cell(row, col2).value))
    
    return ss.spearmanr(x_axis, y_axis).pvalue


for i in range(43, 45):
    for j in range(2, sheet.max_column + 1):
        if i == j:
            continue
        p = float(spvalue(i, j))
        if (p <= 0.05 and p >= 0) or (p >= -0.05 and p <= 0):
            print(str(sheet.cell(1, i).value) + " and " + str(sheet.cell(1, j).value))